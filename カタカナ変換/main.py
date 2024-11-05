import openpyxl
import re
import os
import pickle
import pprint

def make_dict(file_path, brands):
    """Excelファイルから辞書を作成する関数"""
    workbook = openpyxl.load_workbook(file_path)

    reading = {}
    for sheet in brands:
        ws = workbook[sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                word = row[0].strip()  # 単語を取得して空白を除去する
                read = row[1].strip()  # 読みを取得して空白を除去する
                reading[word] = read
    
    #pprint.pprint(reading)
    return reading

def load_dict(file_path, cache_file, brands):
    """辞書をキャッシュからロードする関数"""
    dictionary = None
    updated = False

    if os.path.exists(cache_file):
        with open(cache_file, 'rb') as f:
            cached_data = pickle.load(f)
            # キャッシュの最終更新日時と辞書ファイルの最終更新日時を比較
            if os.path.getmtime(file_path) == cached_data['mtime']:
                dictionary = cached_data['dictionary']
            else:
                updated = True
    else:
        updated = True

    if updated:
        dictionary = make_dict(file_path, brands)
        # 辞書をキャッシュに保存
        with open(cache_file, 'wb') as f:
            pickle.dump({
                'dictionary': dictionary,
                'mtime': os.path.getmtime(file_path)
            }, f)
        print("辞書が更新されました。")
        print("--------------------")

    #pprint.pprint(dictionary)
    return dictionary


def is_alphabetic(word):
    """単語がアルファベットのみで構成されているかどうかを判定する関数"""
    return bool(re.fullmatch(r'[A-Za-z ]+', word))

def replace_text(text, dictionary, not_founds, not_found_else, brand_name):
    """テキスト内のフレーズを辞書を使って変換する関数"""
    words = text.split()
    new_text = []

    i = 0
    while i < len(words):
        current_phrase = ' '.join(words[i:])
        matched_phrase = None

        # 現在の部分を辞書で検索
        for length in range(len(current_phrase), 0, -1):
            subphrase = ' '.join(words[i:i + length])
            if subphrase in dictionary:
                matched_phrase = dictionary[subphrase]
                break
        
        if matched_phrase:
            # 変換が見つかった場合、変換結果を追加し、インデックスを更新
            new_text.append(matched_phrase)
            i += len(subphrase.split())
        else:
            # アルファベットのみの単語で辞書にヒットしなかった場合の処理
            if is_alphabetic(words[i]):
                if brand_name in not_founds.keys():
                    not_founds[brand_name].append(words[i])
                else :
                    not_found_else.append(words[i])
                
            # 変換が見つからなかった場合、そのまま追加
            new_text.append(words[i])
            i += 1

    return ' '.join(new_text)

def update_excel(input_file, output_file, dictionary, brands):
    """Excelファイル内のテキストを辞書を使って変換する関数"""
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active

    not_founds = {}
    not_found_else = []
    for brand in brands:
        not_founds[brand] = []

    # E列の8行目から下のセルを処理
    for row_idx in range(8, sheet.max_row + 1):
        cell = sheet[f'E{row_idx}']
        brand = sheet[f'D{row_idx}']
        if cell.value is not None:
            original_text = str(cell.value)
            #print(f"Original text: {original_text}")

            updated_text = replace_text(original_text, dictionary, not_founds, not_found_else, brand.value)
            #print(f"Updated text: {updated_text}")
            #print("-----------------------------")

            cell.value = updated_text

    workbook.save(output_file)

    # 辞書にない単語があれば最後に表示
    for brand in brands:
        if not_founds[brand]:
            print()
            print(f"{brand}の辞書にない単語:")
            for word in set(not_founds[brand]):  
                print(f"・{word}")
    
    if not_found_else:
        print()
        print("以外の辞書にない単語:")
        for word in set(not_found_else): 
            print(f"・{word}")

    print("------------------")
    print("処理が完了しました。")

def main():
    dict_file = r"C:\Users\lenovo02\Desktop\自動化\カタカナ変換\変換用辞書.xlsx"
    cache_file = r"C:\Users\lenovo02\Desktop\自動化\カタカナ変換\dictionary_cache.pkl"

    ################################################################################
    ################################# 変更部分 ######################################

    brands = ["HERMES", "LOUIS VUITTON", "other"]
    input_file = r"C:\Users\lenovo02\Desktop\自動化\カタカナ変換\コメ兵_原本.xlsm"
    output_file = r"C:\Users\lenovo02\Desktop\自動化\カタカナ変換\test2.xlsx"
    
    ################################################################################
    ################################################################################

    dictionary = load_dict(dict_file, cache_file, brands)
    update_excel(input_file, output_file, dictionary, brands)


if __name__ == "__main__":
    main()