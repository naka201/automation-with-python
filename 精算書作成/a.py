import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
import os
import shutil

# 入力ファイルパス
input_file = "/Users/nakanishitakumi/Library/CloudStorage/OneDrive-KyushuInstituteOfTechnolgy/自動化/automation-with-python/精算書作成/2024-12-ライブ結果.xlsx"
output_file = '/Users/nakanishitakumi/Library/CloudStorage/OneDrive-KyushuInstituteOfTechnolgy/自動化/automation-with-python/精算書作成/集計結果.xlsx'

# 日付の指定（例: '2024-12-01'）
specified_date = None  # ここに日付を指定するか、Noneのままにしておく

# データ読み込み
raw_data = pd.read_excel(input_file, sheet_name='原本')

# 必要な列を抽出
filtered_data = raw_data[['出品者', '通し番号', '販売価格（税込）', '購入者']]

# 「通し番号」のカラム名を変更
filtered_data = filtered_data.rename(columns={'通し番号': '通し'})

# 空の購入者名を除外
filtered_data = filtered_data.dropna(subset=['購入者'])

# 罫線のスタイル
medium_border = Border(
    left=Side(style="medium"),
    right=Side(style="medium"),
    top=Side(style="medium"),
    bottom=Side(style="medium"),
)

# 実行次の日付のフォルダを作成
next_day = (datetime.now() + timedelta(days=1)).strftime('%Y-%m')
folder_path = f"/Users/nakanishitakumi/Library/CloudStorage/OneDrive-KyushuInstituteOfTechnolgy/自動化/automation-with-python/精算書作成/{next_day}"
os.makedirs(folder_path, exist_ok=True)

# 日付を設定
if specified_date:
    date_value = datetime.strptime(specified_date, '%Y-%m-%d')
else:
    date_value = datetime.now()

# 各購入者ごとにファイルを作成
for purchaser, group in filtered_data.groupby('購入者'):
    # ファイルをコピーして購入者名に変更
    purchaser_file = os.path.join(folder_path, f"{purchaser}.xlsx")
    shutil.copy(output_file, purchaser_file)

    # エクセルファイルを読み込み
    wb = load_workbook(purchaser_file)

    # 明細書シートにデータを追加
    if '明細書' in wb.sheetnames:
        sheet = wb['明細書']

        # E3セルに日付を設定
        sheet['E3'] = date_value.strftime('%Y年%m月%d日')
        sheet['E3'].number_format = 'yyyy年mm月dd日'
        sheet['E3'].alignment = Alignment(horizontal="right", vertical="center")

        # 各購入者ごとにデータを追加
        start_row = 6
        start_col = 2  # B列

        # 左側の新しい通し番号を追加
        group = group.reset_index(drop=True)  # 元のインデックスをリセット
        group.insert(0, "通し番号", range(1, len(group) + 1))  # 新しい列を挿入

        # 購入者列を削除
        group = group.drop(columns=['購入者'])

        # データを書き込み
        for r_idx, row in enumerate(dataframe_to_rows(group, index=False, header=True), start=start_row):
            for c_idx, value in enumerate(row, start=start_col):
                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                cell.border = medium_border
                if c_idx == start_col + 3:  # 販売価格列（C列）
                    cell.number_format = '"¥"#,##0'
                if r_idx == start_row:  # ヘッダー行
                    cell.alignment = Alignment(horizontal="center", vertical="center")

        # ヘッダーのC6とD6を結合して「商品番号」に変更
        sheet.merge_cells(start_row=start_row, start_column=start_col+1, end_row=start_row, end_column=start_col+2)
        sheet.cell(row=start_row, column=start_col+1, value="商品番号").alignment = Alignment(horizontal="center", vertical="center")

        start_row += len(group)  # 次の購入者のデータの開始行を設定

        # 2x2の新しい表を追加
        start_row += 1  # 新しい表の開始行（既存の表の下から1行下）

        # 左の列を結合
        sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+2)
        sheet.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+1, end_column=start_col+2)

        # 新しい表の内容を設定
        sheet.cell(row=start_row, column=start_col, value="販売価格合計(税込)10%").alignment = Alignment(horizontal="center", vertical="center")
        cell_sum = sheet.cell(row=start_row, column=start_col+3, value='=SUM(E7:E{})'.format(start_row-1))
        sheet.cell(row=start_row+1, column=start_col, value="内消費税金額(10％)").alignment = Alignment(horizontal="center", vertical="center")
        cell_tax = sheet.cell(row=start_row+1, column=start_col+3, value='=E{}*0.1'.format(start_row))

        # 通貨書式を適用
        cell_sum.number_format = '"¥"#,##0'
        cell_tax.number_format = '"¥"#,##0'

        # 罫線を適用
        for row in sheet.iter_rows(min_row=start_row, max_row=start_row+1, min_col=start_col, max_col=start_col+3):
            for cell in row:
                cell.border = medium_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
    else:
        raise ValueError("明細書シートが見つかりません")

    # 保存し直す
    wb.save(purchaser_file)

print(f"購入者ごとのデータが {folder_path} に保存されました（ヘッダー統合、罫線、通貨書式を適用）。")