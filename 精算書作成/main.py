import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
import os
import shutil

def generate_settlement_files(input_file, output_file, file_cnt, specified_date=None):
    """
    指定されたExcelファイルを基に購入者ごとの精算書を作成する関数。
    """
    # データ読み込み
    raw_data = pd.read_excel(input_file, sheet_name='2025.01')

    # 罫線のスタイル
    medium_border = Border(
        left=Side(style="medium"),
        right=Side(style="medium"),
        top=Side(style="medium"),
        bottom=Side(style="medium"),
    )

    # 実行次の日付のフォルダを作成
    next_day = (datetime.now() + timedelta(days=1)).strftime('%Y-%m')
    folder_path = f"/Users/nakanishitakumi/Library/CloudStorage/OneDrive-KyushuInstituteOfTechnolgy/自動化/automation-with-python/精算書作成/{next_day}"
    os.makedirs(folder_path, exist_ok=True)

    # 日付を設定
    if specified_date:
        date_value = datetime.strptime(specified_date, '%Y-%m-%d')
    else:
        date_value = datetime.now()

    if "販売" in output_file:
        # 必要な列を抽出
        filtered_data = raw_data[['出品者', '通し番号', 'LIVE小計（税込）', '購入者']]

        # 「通し番号」のカラム名を変更
        filtered_data = filtered_data.rename(columns={'通し番号': '通し', 'LIVE小計（税込）': '販売価格'})

        # 空の購入者名を除外
        filtered_data = filtered_data.dropna(subset=['購入者'])

        # 購入者ごとの手数料（税込）の合計を計算
        fee_data = raw_data[['出品者', 'LIVE手数料（税込）', '購入者']]
        fee_data = fee_data.dropna(subset=['購入者'])
        fees_dict = fee_data.groupby('購入者')['LIVE手数料（税込）'].sum().to_dict()

        # 各購入者ごとにファイルを作成
        for purchaser, group in filtered_data.groupby('購入者'):
            # ファイルをコピーして購入者名に変更
            purchaser_file = os.path.join(folder_path, f"{purchaser}.xlsx")
            shutil.copy(output_file, purchaser_file)

            # エクセルファイルを読み込み
            wb = load_workbook(purchaser_file)

            # 明細書シートにデータを追加
            if '明細書' in wb.sheetnames:
                sheet = wb['明細書']

                # E3セルに日付を設定
                sheet['E3'] = date_value.strftime('%Y年%m月%d日')
                sheet['E3'].number_format = 'yyyy年mm月dd日'
                sheet['E3'].alignment = Alignment(horizontal="right", vertical="center")

                # 各購入者ごとにデータを追加
                start_row = 6
                start_col = 2  # B列

                # 左側の新しい通し番号を追加
                group = group.reset_index(drop=True)  # 元のインデックスをリセット
                group.insert(0, "通し番号", range(1, len(group) + 1))  # 新しい列を挿入

                # 購入者列を削除
                group = group.drop(columns=['購入者'])

                # データを書き込み
                for r_idx, row in enumerate(dataframe_to_rows(group, index=False, header=True), start=start_row):
                    for c_idx, value in enumerate(row, start=start_col):
                        cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                        cell.border = medium_border
                        if c_idx == start_col + 3:  # 販売価格列（C列）
                            cell.number_format = '"¥"#,##0'
                        if r_idx == start_row:  # ヘッダー行
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                # ヘッダーのC6とD6を結合して「商品番号」に変更
                sheet.merge_cells(start_row=start_row, start_column=start_col+1, end_row=start_row, end_column=start_col+2)
                sheet.cell(row=start_row, column=start_col+1, value="商品番号").alignment = Alignment(horizontal="center", vertical="center")

                start_row += len(group)  # 次の購入者のデータの開始行を設定

                # 2x2の新しい表を追加
                start_row += 1  # 新しい表の開始行（既存の表の下から1行下）

                # 左の列を結合
                sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+2)
                sheet.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+1, end_column=start_col+2)

                # 新しい表の内容を設定
                sheet.cell(row=start_row, column=start_col, value="販売価格合計(税込)10%").alignment = Alignment(horizontal="center", vertical="center")
                cell_sum = sheet.cell(row=start_row, column=start_col+3, value='=SUM(E6:E{})'.format(start_row-1))
                sheet.cell(row=start_row+1, column=start_col, value="内消費税金額(10％)").alignment = Alignment(horizontal="center", vertical="center")
                cell_tax = sheet.cell(row=start_row+1, column=start_col+3, value='=E{}*0.1'.format(start_row))

                # 通貨書式を適用
                cell_sum.number_format = '"¥"#,##0'
                cell_tax.number_format = '"¥"#,##0'

                # 罫線を適用
                for row in sheet.iter_rows(min_row=start_row, max_row=start_row+1, min_col=start_col, max_col=start_col+3):
                    for cell in row:
                        cell.border = medium_border

                # 手数料の表
                start_row += 2  # 新しい表の開始行（既存の表の下から1行下）

                # 左の列を結合
                sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+2)
                sheet.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+1, end_column=start_col+2)

                # 新しい表の内容を設定
                sheet.cell(row=start_row, column=start_col, value="手数料合計(税込)10%").alignment = Alignment(horizontal="center", vertical="center")
                cell_sum = sheet.cell(row=start_row, column=start_col+3, value=fees_dict[purchaser])
                sheet.cell(row=start_row+1, column=start_col, value="内消費税金額(10％)").alignment = Alignment(horizontal="center", vertical="center")
                cell_tax = sheet.cell(row=start_row+1, column=start_col+3, value='=E{}*0.1'.format(start_row))

                # 通貨書式を適用
                cell_sum.number_format = '"¥"#,##0'
                cell_tax.number_format = '"¥"#,##0'

                # 罫線を適用
                for row in sheet.iter_rows(min_row=start_row, max_row=start_row+1, min_col=start_col, max_col=start_col+3):
                    for cell in row:
                        cell.border = medium_border

                # 集計データの計算
                start_row_data = 6  # データの開始行
                end_row_data = start_row - 1  # データの終了行
                count = end_row_data - start_row_data

                # 合計計算
                sum_value = sum(
                    sheet.cell(row=r, column=start_col+3).value
                    for r in range(start_row_data, end_row_data + 1)
                    if isinstance(sheet.cell(row=r, column=start_col+3).value, (int, float))
                )

                # 精算書_販売シートの更新
                settlement_sheet = wb['精算書_販売']
                settlement_sheet["A20"] = date_value.strftime('%m/%d')
                settlement_sheet["A5"] = purchaser
                settlement_sheet["M20"] = count
                settlement_sheet["O20"] = sum_value  

                # 保存し直す
                wb.save(purchaser_file)

                print(f"{purchaser}の処理が完了しました。")

    elif "出品" in output_file:
        # 必要な列を抽出
        filtered_data = raw_data[['出品者', '通し番号', '販売価格（税込）', '購入者']]

        # 「通し番号」のカラム名を変更
        filtered_data = filtered_data.rename(columns={'通し番号': '通し', '販売価格（税込）': '販売価格'})

        # 空の購入者名を除外
        filtered_data = filtered_data.dropna(subset=['購入者'])

        # データ読み込み
        seller_data = pd.read_excel(input_file, sheet_name='出品者対応表')
        # 出品者と出品者名の辞書を作成
        sellers_dict = dict(zip(seller_data['出品者'], seller_data['出品者名']))
        sells_dict = filtered_data.groupby('出品者')['販売価格'].sum().to_dict()

        # 出品者ごとの手数料（税込）の合計を計算
        fee_data = raw_data[['出品者', '手数料（税込）', '購入者']]
        fee_data = fee_data.dropna(subset=['購入者'])
        fees_dict = fee_data.groupby('出品者')['手数料（税込）'].sum().to_dict()

        # 各出品者ごとにファイルを作成
        for seller, group in filtered_data.groupby('出品者'):
            # ファイルをコピーして出品者名に変更
            seller_file = os.path.join(folder_path, f"{seller}.xlsx")
            shutil.copy(output_file, seller_file)

            # エクセルファイルを読み込み
            wb = load_workbook(seller_file)

            # 明細書シートにデータを追加
            if '明細書' in wb.sheetnames:
                sheet = wb['明細書']

                # E3セルに日付を設定
                sheet['E3'] = date_value.strftime('%Y年%m月%d日')
                sheet['E3'].number_format = 'yyyy年mm月dd日'
                sheet['E3'].alignment = Alignment(horizontal="right", vertical="center")

                # 各購入者ごとにデータを追加
                start_row = 6
                start_col = 2  # B列

                # 左側の新しい通し番号を追加
                group = group.reset_index(drop=True)  # 元のインデックスをリセット
                group.insert(0, "通し番号", range(1, len(group) + 1))  # 新しい列を挿入

                # 購入者列を削除
                group = group.drop(columns=['購入者'])

                # データを書き込み
                for r_idx, row in enumerate(dataframe_to_rows(group, index=False, header=True), start=start_row):
                    for c_idx, value in enumerate(row, start=start_col):
                        cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                        cell.border = medium_border
                        if c_idx == start_col + 3:  # 販売価格列（C列）
                            cell.number_format = '"¥"#,##0'
                        if r_idx == start_row:  # ヘッダー行
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                # ヘッダーのC6とD6を結合して「商品番号」に変更
                sheet.merge_cells(start_row=start_row, start_column=start_col+1, end_row=start_row, end_column=start_col+2)
                sheet.cell(row=start_row, column=start_col+1, value="商品番号").alignment = Alignment(horizontal="center", vertical="center")

                start_row += len(group)  # 次の購入者のデータの開始行を設定

                # 価格の表
                start_row += 1  # 新しい表の開始行（既存の表の下から1行下）

                # 左の列を結合
                sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+2)
                sheet.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+1, end_column=start_col+2)

                # 新しい表の内容を設定
                sheet.cell(row=start_row, column=start_col, value="販売価格合計(税込)10%").alignment = Alignment(horizontal="center", vertical="center")
                cell_sum = sheet.cell(row=start_row, column=start_col+3, value='=SUM(E6:E{})'.format(start_row-1))
                sheet.cell(row=start_row+1, column=start_col, value="内消費税金額(10％)").alignment = Alignment(horizontal="center", vertical="center")
                cell_tax = sheet.cell(row=start_row+1, column=start_col+3, value='=E{}/(1+0.1)*0.1'.format(start_row))

                # 通貨書式を適用
                cell_sum.number_format = '"¥"#,##0'
                cell_tax.number_format = '"¥"#,##0'

                # 罫線を適用
                for row in sheet.iter_rows(min_row=start_row, max_row=start_row+1, min_col=start_col, max_col=start_col+3):
                    for cell in row:
                        cell.border = medium_border

                # 手数料の表
                start_row += 2  # 新しい表の開始行（既存の表の下から1行下）

                # 左の列を結合
                sheet.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col+2)
                sheet.merge_cells(start_row=start_row+1, start_column=start_col, end_row=start_row+1, end_column=start_col+2)

                # 新しい表の内容を設定
                sheet.cell(row=start_row, column=start_col, value="手数料合計(税込)10%").alignment = Alignment(horizontal="center", vertical="center")
                cell_sum = sheet.cell(row=start_row, column=start_col+3, value=fees_dict[seller])
                sheet.cell(row=start_row+1, column=start_col, value="内消費税金額(10％)").alignment = Alignment(horizontal="center", vertical="center")
                cell_tax = sheet.cell(row=start_row+1, column=start_col+3, value='=E{}/(1+0.1)*0.1'.format(start_row))

                # 通貨書式を適用
                cell_sum.number_format = '"¥"#,##0'
                cell_tax.number_format = '"¥"#,##0'

                # 罫線を適用
                for row in sheet.iter_rows(min_row=start_row, max_row=start_row+1, min_col=start_col, max_col=start_col+3):
                    for cell in row:
                        cell.border = medium_border

                # 集計データの計算
                start_row_data = 6  # データの開始行
                end_row_data = start_row - 1  # データの終了行
                count = end_row_data - start_row_data

                # 合計計算
                sum_value = sum(
                    sheet.cell(row=r, column=start_col+3).value
                    for r in range(start_row_data, end_row_data + 1)
                    if isinstance(sheet.cell(row=r, column=start_col+3).value, (int, float))
                )

                # 精算書_販売シートの更新
                settlement_sheet = wb['精算書_出品']
                settlement_sheet["A19"] = date_value.strftime('%m/%d')
                settlement_sheet["A20"] = date_value.strftime('%m/%d')
                settlement_sheet["A24"] = date_value.strftime('%m/%d')
                settlement_sheet["A5"] = sellers_dict[seller]
                settlement_sheet["M24"] = count
                settlement_sheet["O19"] = fees_dict[seller]
                settlement_sheet["O24"] = sum_value 
                settlement_sheet["O25"] = sells_dict[seller]

                if sum_value != sells_dict[seller]:
                    print(f"{sellers_dict[seller]}のデータに誤りがあります。") 

                # 保存し直す
                wb.save(seller_file)

                print(f"{sellers_dict[seller]}の処理が完了しました。")

    print(f"全ての処理が完了しました。")


# 使用例
input_file = "/Users/nakanishitakumi/Library/CloudStorage/OneDrive-KyushuInstituteOfTechnolgy/自動化/automation-with-python/精算書作成/2025.01.07.xlsx"
settlement_buy = '/Users/nakanishitakumi/Library/CloudStorage/OneDrive-KyushuInstituteOfTechnolgy/自動化/automation-with-python/精算書作成/精算書_販売.xlsx'
settlement_sell = '/Users/nakanishitakumi/Library/CloudStorage/OneDrive-KyushuInstituteOfTechnolgy/自動化/automation-with-python/精算書作成/精算書_出品.xlsx'

file_cnt = 0
specified_date = "2024-12-07"

#generate_settlement_files(input_file, settlement_buy)
generate_settlement_files(input_file, settlement_sell)