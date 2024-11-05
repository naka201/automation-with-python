import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import time

#ファイル名から数字を抽出する
def num_get(file_path):
    filename = os.path.basename(file_path)
    match = re.match(r'(\d{1,4})', filename)
    return match.group(1) if match else None

#データをペーストする
def paste_data(file_path, pasted_file):
    num = num_get(file_path)
    print(f"業者コード: {num}")

    df1 = pd.read_excel(file_path, header=None)
    df2 = pd.read_excel(pasted_file, header=None)
    
    first_row = df2.iloc[0].astype(str).tolist()
    try:
        column_idx = first_row.index(num)
        
        # df1の3列目から値を抽出
        values_to_paste = df1[3]

        required_rows = 4 + len(values_to_paste)

        new_rows = pd.DataFrame([[None]*df2.shape[1]] * (required_rows - len(df2)), columns=df2.columns)
        df2 = pd.concat([df2, new_rows], ignore_index=True)
        
        for i, value in enumerate(values_to_paste):
            df2.iloc[i, column_idx] = value if pd.notna(value) else ''
    except ValueError:
        print("数字が見つかりませんでした。")
        exit()

    df2.to_excel(pasted_file, index=False, header=False)
    
    # openpyxlを使って黄色で塗りつぶし
    wb = load_workbook(pasted_file)
    ws = wb.active

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for i in range(3, len(values_to_paste)):
        value = values_to_paste.iloc[i]
        cell = ws.cell(row=i + 1, column=column_idx + 1)  
        if isinstance(value, str):
            cell.fill = yellow_fill

    wb.save(pasted_file)

# dataフォルダ内のすべての.xlsxファイルをリストアップする関数
def list_excel_files(folder_path):
    return [os.path.join(folder_path, file) for file in os.listdir(folder_path) if file.endswith('.xlsx')]

def paste(folder_path, pasted_file):
    print(f"{folder_path}フォルダ内のデータの処理を開始します。")
    print("-----")

    paste_files = list_excel_files(folder_path)

    start_time = time.time()  # 開始時刻を記録

    for file_path in paste_files:
        print(f"Processing {file_path}...")
        if not os.path.exists(file_path):
            print(f"{file_path}が見つかりません。スキップします。")
            continue
        paste_data(file_path, pasted_file)
        print(f"{file_path} のデータをペーストしました。")
        print("-----")

    end_time = time.time()  # 終了時刻を記録
    elapsed_time = end_time - start_time
    print("全てのファイルの処理が完了しました。")
    print(f"処理時間: {elapsed_time:.2f}秒")


if __name__ == "__main__":
    #################################################################################
    ####################### 使用時変更部分 ###########################################

    # フォルダ内のすべての.xlsxファイルを取得
    # バッグ・時計・宝石で使用しないものに、先頭に"#"を付けて下さい。

    #folder_path = r'C:\Users\lenovo02\Desktop\自動化\入札貼付\バッグ用データ'
    #pasted_file = r'C:\Users\lenovo02\Desktop\自動化\入札貼付\バッグ_貼り付け先.xlsx'

    #folder_path = r'C:\Users\lenovo02\Desktop\自動化\入札貼付\時計用データ'
    #pasted_file = r'C:\Users\lenovo02\Desktop\自動化\入札貼付\時計_貼り付け先.xlsx'

    folder_path = r'C:\Users\lenovo02\Desktop\自動化\入札貼付\宝石用データ'
    pasted_file = r'C:\Users\lenovo02\Desktop\自動化\入札貼付\宝石_貼り付け先.xlsx'

    #################################################################################
    #################################################################################
    
    # メイン関数の実行
    paste(folder_path, pasted_file)