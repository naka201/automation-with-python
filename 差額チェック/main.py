import pandas as pd
import time
import openpyxl
from openpyxl.styles import PatternFill

def bag_dataframe(file_path):
    """
    Excelファイルからデータフレームを作成し、不要な行と列を削除する。
    """
    df = pd.read_excel(file_path, engine='openpyxl')
    df = df.drop([0, 1]).reset_index(drop=True)  # 2行目と3行目を削除
    df = df.drop(columns=['メモ', '詳細'], errors='ignore')  # 不要な列を削除
    df_main = df.iloc[:, :10]  # L列目までを抽出
    return df_main

def watch_dataframe(file_path):
    """
    Excelファイルからデータフレームを作成し、不要な行と列を削除する。
    """
    df = pd.read_excel(file_path, engine='openpyxl')
    df = df.drop([0]).reset_index(drop=True)  # 2行目と3行目を削除
    df = df.drop(columns=['メモ', '詳細'], errors='ignore')  # 不要な列を削除
    df_main = df.iloc[:, :10]  # L列目までを抽出
    return df_main

def jewel_dataframe(file_path):
    """
    Excelファイルからデータフレームを作成し、不要な行と列を削除する。
    """
    df = pd.read_excel(file_path, engine='openpyxl')
    df = df.drop([0]).reset_index(drop=True)  # 2行目と3行目を削除
    df = df.drop(columns=['メモ', '詳細'], errors='ignore')  # 不要な列を削除
    df_main = df.iloc[:, :11]  # L列目までを抽出
    return df_main

def bag_filter(df):
    """
    条件:
    - MAX金額：100万以上、差額：50万以上
    - MAX金額：50万以上100万未満、差額：20万以上
    - MAX金額：30万以上50万未満、差額：10万以上
    - MAX金額：10万以上30万未満、差額：5万以上
    - MAX金額：1万以上10万未満、差額：3万以上
    - MAX金額：1万以下、差額：9000円以上
    """
    # 各条件に基づいてフィルタリング
    filtered = df[
        ((df['MAX金額'] >= 1000000) & (df['差額'] >= 500000)) |
        ((df['MAX金額'] >= 500000) & (df['MAX金額'] < 1000000) & (df['差額'] >= 200000)) |
        ((df['MAX金額'] >= 300000) & (df['MAX金額'] < 500000) & (df['差額'] >= 100000)) |
        ((df['MAX金額'] >= 100000) & (df['MAX金額'] < 300000) & (df['差額'] >= 50000)) |
        ((df['MAX金額'] >= 10000) & (df['MAX金額'] < 100000) & (df['差額'] >= 30000)) |
        ((df['MAX金額'] <= 10000) & (df['差額'] >= 9000))
    ]
    
    # フィルタリングした行のインデックスをリストに変換
    indexes = [index + 4 for index in filtered.index.tolist()]
    
    return indexes

def watch_filter(df):
    """
    条件:
    - MAX金額：100万以上、差額：20万以上
    - MAX金額：10万以上100万未満、差額：10万以上
    - MAX金額：1万以上10万未満、差額：5万以上
    - MAX金額：1円以上1万未満、差額：5000円以上
    """
    # 各条件に基づいてフィルタリング
    filtered = df[
        ((df['MAX金額'] >= 1000000) & (df['差額'] >= 200000)) |
        ((df['MAX金額'] >= 100000) & (df['MAX金額'] < 1000000) & (df['差額'] >= 100000)) |
        ((df['MAX金額'] >= 10000) & (df['MAX金額'] < 100000) & (df['差額'] >= 50000)) |
        ((df['MAX金額'] <= 1) & (df['MAX金額'] < 10000) & (df['差額'] >= 5000)) 
    ]
    
    # フィルタリングした行のインデックスをリストに変換
    indexes = [index + 3 for index in filtered.index.tolist()]
    
    return indexes

def jewel_filter(df):
    """
    条件:
    - MAX金額：50万以上、差額：20万以上
    - MAX金額：30万以上50万未満、差額：10万以上
    - MAX金額：10万以上30万未満、差額：3万以上
    - MAX金額：1万以上10万未満、差額：1万以上
    - MAX金額：1円以上1万未満、差額：3500円以上
    """
    # 各条件に基づいてフィルタリング
    filtered = df[
        ((df['MAX金額'] >= 500000) & (df['差額'] >= 200000)) |
        ((df['MAX金額'] >= 300000) & (df['MAX金額'] < 500000) & (df['差額'] >= 100000)) |
        ((df['MAX金額'] >= 100000) & (df['MAX金額'] < 300000) & (df['差額'] >= 30000)) |
        ((df['MAX金額'] >= 10000) & (df['MAX金額'] < 100000) & (df['差額'] >= 10000)) |
        ((df['MAX金額'] <= 1) & (df['MAX金額'] < 10000) & (df['差額'] >= 3500))
    ]
    
    # フィルタリングした行のインデックスをリストに変換
    indexes = [index + 3 for index in filtered.index.tolist()]
    
    return indexes

def reset_column_color(sheet, column):
    """
    12列目の全てのセルを白色にリセットする。
    """
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=column, max_col=column):
        for cell in row:
            cell.fill = white_fill

def bag_coloring(file_path, indexes):
    """
    ExcelファイルのL列に指定された行のセルに黄色の背景色を塗る。
    """
    # Excelファイルの読み込み
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # 黄色の背景色を設定するためのパターンフィル
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # L列を白色にリセット
    reset_column_color(sheet, column=12)

    # 指定された行に黄色の背景色を塗る
    for row in indexes:
        cell = sheet.cell(row=row, column=12) 
        cell.fill = yellow_fill

    # Excelファイルを保存する
    output_file_path = file_path.replace(".xlsx", "_チェック後.xlsx")
    workbook.save(output_file_path)
    print()
    print(f"条件を満たすセルに背景色を塗りました。")

def watch_coloring(file_path, indexes):
    """
    ExcelファイルのL列に指定された行のセルに黄色の背景色を塗る。
    """
    # Excelファイルの読み込み
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # 黄色の背景色を設定するためのパターンフィル
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # L列を白色にリセット
    reset_column_color(sheet, column=10)

    # 指定された行に黄色の背景色を塗る
    for row in indexes:
        cell = sheet.cell(row=row, column=10) 
        cell.fill = yellow_fill

    # Excelファイルを保存する
    output_file_path = file_path.replace(".xlsx", "_チェック後.xlsx")
    workbook.save(output_file_path)
    print()
    print(f"条件を満たすセルに背景色を塗りました。")

def jewel_coloring(file_path, indexes):
    """
    ExcelファイルのL列に指定された行のセルに黄色の背景色を塗る。
    """
    # Excelファイルの読み込み
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # 黄色の背景色を設定するためのパターンフィル
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # L列を白色にリセット
    reset_column_color(sheet, column=11)

    # 指定された行に黄色の背景色を塗る
    for row in indexes:
        cell = sheet.cell(row=row, column=11) 
        cell.fill = yellow_fill

    # Excelファイルを保存する
    output_file_path = file_path.replace(".xlsx", "_チェック後.xlsx")
    workbook.save(output_file_path)
    print()
    print(f"条件を満たすセルに背景色を塗りました。")

def main():
    #使用時変更部分 
    file_path = r"C:\Users\lenovo02\Desktop\自動化\差額チェック\【第92回バッグ大会】集計表シート_原本.xlsx"
    
    # 処理開始時間を記録
    start_time = time.time()

    # バッグ大会
    if "バッグ" in file_path:
        print("-----------------------------------")
        print("バッグ大会の差額チェックを行います")
        print("-----------------------------------")
        
        bag_df = bag_dataframe(file_path)
        print(bag_df.head())

        indexes = bag_filter(bag_df)
        bag_coloring(file_path, indexes)

    # 時計大会
    elif "時計" in file_path:
        print("---------------------------------")
        print("時計大会の差額チェックを行います")
        print("---------------------------------")

        watch_df = watch_dataframe(file_path)
        print(watch_df.head())

        indexes = watch_filter(watch_df)
        watch_coloring(file_path, indexes)

    # 宝石大会
    elif "宝石" in file_path:
        print("---------------------------------")
        print("宝石大会の差額チェックを行います")
        print("---------------------------------")

        jewel_df = jewel_dataframe(file_path)
        print(jewel_df.head())

        indexes = jewel_filter(jewel_df)
        jewel_coloring(file_path, indexes)
    

    # 処理終了時間を記録
    end_time = time.time()

    # 実行時間を計算
    execution_time = end_time - start_time

    print(f"条件を満たす行数: {len(indexes)}")
    print(f"処理時間: {execution_time:.2f}秒")
    

if __name__ == "__main__":
    main()
