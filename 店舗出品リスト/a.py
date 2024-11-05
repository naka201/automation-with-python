import shutil
import openpyxl
import os

def copy_cells(src_file, src_sheet, src_range, dest_file, dest_sheet, dest_start_cell):
    # 元のExcelファイルを開く
    src_wb = openpyxl.load_workbook(src_file)

    # 元のシートが存在するか確認
    if src_sheet not in src_wb.sheetnames:
        raise KeyError(f"元のシート {src_sheet} が存在しません。")

    src_ws = src_wb[src_sheet]

    # コピー先のExcelファイルを開く（存在しない場合は新規作成）
    try:
        dest_wb = openpyxl.load_workbook(dest_file)
        if dest_sheet not in dest_wb.sheetnames:
            dest_ws = dest_wb.create_sheet(dest_sheet)
        else:
            dest_ws = dest_wb[dest_sheet]
    except FileNotFoundError:
        dest_wb = openpyxl.Workbook()
        dest_ws = dest_wb.active
        dest_ws.title = dest_sheet

    # セル範囲を取得
    src_cells = src_ws[src_range]

    # 貼り付け先の開始セルを指定
    dest_start_row, dest_start_col = openpyxl.utils.cell.coordinate_to_tuple(dest_start_cell)
        
    # セルの内容をコピー
    if isinstance(src_cells, openpyxl.cell.cell.Cell):
        # 単一セルの場合
        dest_cell = dest_ws.cell(row=dest_start_row, column=dest_start_col)
        existing_value = dest_cell.value if dest_cell.value else ""
        new_value = src_cells.value
        if isinstance(new_value, str):
            new_value = new_value.replace('市場', '')  # 市場という文字列を除去
        if new_value:
            if existing_value:
                if isinstance(new_value, float):
                    new_value = int(new_value)
                dest_cell.value = f"{existing_value}, {new_value}" if isinstance(new_value, str) else new_value
            else:
                dest_cell.value = new_value

        #print(f"貼り付け先セル {dest_start_cell}: {dest_cell.value}")  # デバッグ用出力
    else:
        # 複数セルの場合
        for i, row in enumerate(src_cells):
            for j, cell in enumerate(row):
                dest_row = dest_start_row + i
                dest_col = dest_start_col + j
                dest_cell = dest_ws.cell(row=dest_row, column=dest_col)
                existing_value = dest_cell.value if dest_cell.value else ""
                new_value = cell.value
                if isinstance(new_value, str):
                    new_value = new_value.replace('市場', '')  # 市場という文字列を除去
                if new_value:
                    col = openpyxl.utils.cell.get_column_letter(dest_cell.column)
                    if col == "D":
                        val = new_value.split()
                        if val[0].isdigit():
                            new_value = val[1]

                    if existing_value:
                        if isinstance(new_value, float):
                            new_value = int(new_value)
                        dest_cell.value = f"{existing_value}, {new_value}" if isinstance(new_value, str) or isinstance(new_value, int) else new_value
                    else:
                        dest_cell.value = new_value
                #print(f"貼り付け先セル {dest_start_cell}: {dest_cell.value}")  # デバッグ用出力

    # コピー先のExcelファイルを保存
    dest_wb.save(dest_file)

def copy_excel_file(src_file, dest_file):
    """
    Excelファイルをコピーして新しいファイルを作成する
    """
    shutil.copy(src_file, dest_file)

def get_start_end_for_value(ws, value):
    """
    指定した値が最初に現れる行と最後に現れる行を取得
    """
    start_row = None
    end_row = None
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == value:
            if start_row is None:
                start_row = row_idx
            end_row = row_idx
    
    if start_row is None or end_row is None:
        # 値が見つからない場合は None を返す
        return None, None
    
    start_row, end_row = int(start_row), int(end_row)
    return start_row, end_row

# メイン処理
src_file = r"C:\Users\lenovo02\Desktop\自動化\店舗出品リスト\【原本】店舗出品リスト.xlsx" # 基本的に変更しない
src_sheet = '2024.07'
dest_sheet = '出品表'

copy_file = r"C:\Users\lenovo02\Desktop\自動化\店舗出品リスト\バッグ出品表.xlsx"
output_dir = r"C:\Users\lenovo02\Desktop\自動化\店舗出品リスト\output"

# 出力ディレクトリが存在しない場合は作成
os.makedirs(output_dir, exist_ok=True)

# ファイルXを読み込み
wb_x = openpyxl.load_workbook(src_file)
ws_x = wb_x.active

# A列のユニークな値を取得
box_nums = set()
for cell in ws_x['A']:
    if cell.row > 2:  # ヘッダー行をスキップ
        if cell.value != None:
            num = int(cell.value)
            box_nums.add(num)
box_nums = sorted(box_nums)

# 各ユニークな値ごとに処理
for value in box_nums:
    #print(f"{value}")
    # 箱番のstartとendを取得
    start, end = get_start_end_for_value(ws_x, value)
    #print(f"start = {start}, end = {end}")
    
    if start is None or end is None:
        continue  # 箱番が見つからない場合はスキップ
    
    # 新しいファイルZのパスを設定
    dest_file = os.path.join(output_dir, f'{value}_バッグ出品表.xlsx')
    
    # 新しいExcelファイルZを作成
    copy_excel_file(copy_file, dest_file)

    # 貼り付け処理を開始
    print(f"貼り付けを開始します (箱番: {value})")

    # 箱番
    src_range = f'A{start}'         
    dest_start_cell = 'C4'    
    copy_cells(src_file, src_sheet, src_range, dest_file, dest_sheet, dest_start_cell)

    # 品名
    src_range = f'B{start}:B{end}'  
    dest_start_cell = 'C9'
    copy_cells(src_file, src_sheet, src_range, dest_file, dest_sheet, dest_start_cell)

    # ブランド名
    src_range = f'N{start}:N{end}'
    dest_start_cell = 'D9'
    copy_cells(src_file, src_sheet, src_range, dest_file, dest_sheet, dest_start_cell)

    # 商品名
    src_range = f'H{start}:H{end}'
    dest_start_cell = 'E9'
    copy_cells(src_file, src_sheet, src_range, dest_file, dest_sheet, dest_start_cell)

    # 型番
    src_range = f'I{start}:I{end}' 
    copy_cells(src_file, src_sheet, src_range, dest_file, dest_sheet, dest_start_cell)

    # 付属品
    src_range = f'AB{start}:AB{end}'
    dest_start_cell = 'G9'
    copy_cells(src_file, src_sheet, src_range, dest_file, dest_sheet, dest_start_cell)

    # 指値
    src_range = f'AI{start}:AI{end}' 
    dest_start_cell = 'H9'
    copy_cells(src_file, src_sheet, src_range, dest_file, dest_sheet, dest_start_cell)

    print(f"貼り付けが完了しました (箱番: {value})")

print("すべての処理が完了しました。")
