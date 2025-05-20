# F-project/app/services/brand_search_service.py
import pdfplumber
import pandas as pd
import openpyxl
from openpyxl.styles import Font
import os
from collections import defaultdict
import mojimoji
import tempfile
import shutil
from io import BytesIO # BytesIOをインポート

# --- read_excel_to_set 関数の修正 ---
def read_excel_to_set(excel_bytes_io, column_index=2, start_row=2): # 引数名を変更
    """
    ExcelのBytesIOオブジェクトを読み込み、指定された列の値をセットとして返す。
    """
    brand_set = set()
    try:
        # BytesIOオブジェクトを直接openpyxlに渡す
        workbook = openpyxl.load_workbook(excel_bytes_io)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=start_row, min_col=column_index, max_col=column_index, values_only=True):
            cell_value = row[0]
            if cell_value is not None:
                brand_name = str(cell_value).strip()
                if brand_name:
                    normalized_brand = mojimoji.han_to_zen(brand_name, kana=True)
                    brand_set.add(normalized_brand)
    except Exception as e:
        print(f"Excelファイルの読み込み中にエラーが発生しました: {e}")
        raise
    return brand_set

# --- find_brands_in_pdf_plumber 関数の修正 ---
def find_brands_in_pdf_plumber(pdf_bytes_io, brand_set, target_column_header="ブランド/モデル/型番/重量/状態等", target_column_index=2): # 引数名を変更
    """
    pdfplumberを使用してPDFのBytesIOオブジェクトを読み込み、ブランド名を検索する。
    """
    found_brands_by_page = defaultdict(set)
    print(f"PDF処理を開始します...")
    try:
        # BytesIOオブジェクトを直接pdfplumberに渡す
        with pdfplumber.open(pdf_bytes_io) as pdf:
            total_pages = len(pdf.pages)
            print(f"PDFの全ページ数: {total_pages}")

            # ... (以降のPDF処理ロジックは変更なし) ...
            for page_num_zero_based, page in enumerate(pdf.pages):
                page_num = page_num_zero_based + 1
                try:
                    table_settings = {
                        "vertical_strategy": "lines", "horizontal_strategy": "lines",
                        "snap_tolerance": 3, "join_tolerance": 3,
                    }
                    tables = page.extract_tables(table_settings=table_settings)
                    if not tables:
                        table_settings_text = {
                           "vertical_strategy": "text", "horizontal_strategy": "text",
                           "snap_tolerance": 5, "join_tolerance": 5,
                        }
                        tables = page.extract_tables(table_settings=table_settings_text)

                    if not tables: continue

                    for i, table_data in enumerate(tables):
                        if not table_data: continue

                        df = None
                        header = table_data[0]
                        try:
                            if header and all(isinstance(h, (str, type(None))) for h in header):
                                clean_header = [f"col_{j}" if h is None else str(h).replace('\n', ' ') for j, h in enumerate(header)]
                                df = pd.DataFrame(table_data[1:], columns=clean_header)
                            else:
                                df = pd.DataFrame(table_data)
                        except Exception:
                             continue

                        if df.empty: continue

                        actual_col_ref = None
                        potential_cols_header = [col for col in df.columns if col is not None and target_column_header.split('/')[0] in str(col)]

                        if potential_cols_header:
                            actual_col_ref = potential_cols_header[0]
                        elif len(df.columns) > target_column_index:
                            actual_col_ref = target_column_index
                        else:
                            continue
                        
                        try:
                            if isinstance(actual_col_ref, str):
                                if actual_col_ref not in df.columns:
                                    continue
                                target_column_data = df[actual_col_ref]
                            elif isinstance(actual_col_ref, int):
                                if actual_col_ref >= len(df.columns):
                                    continue
                                target_column_data = df.iloc[:, actual_col_ref]
                            else:
                                continue

                            for cell_value in target_column_data.dropna():
                                cell_text_raw = str(cell_value).strip().replace('\n', ' ')
                                if not cell_text_raw: continue
                                cell_text_normalized = mojimoji.han_to_zen(cell_text_raw, kana=True)
                                for brand in brand_set:
                                    if brand in cell_text_normalized:
                                        found_brands_by_page[page_num].add(brand)
                        except Exception as search_err:
                            print(f"エラー: ページ {page_num}, テーブル {i+1}, 列 '{actual_col_ref}' の検索中にエラー: {search_err}")
                except Exception as page_proc_err:
                    print(f"エラー: ページ {page_num} のテーブル抽出/処理中にエラー: {page_proc_err}")

    except Exception as e:
        print(f"PDFファイルのオープンまたは処理中に予期せぬエラーが発生しました: {e}")
        raise
    return dict(found_brands_by_page)

# --- highlight_matching_cells 関数の修正 ---
def highlight_matching_cells(target_excel_bytes_io, result_dict, sheet_name, output_path): # 引数名を変更
    """
    指定したExcelのBytesIOオブジェクトの中で、result_dictのキーに一致するセルの文字色を変更し、
    output_pathに保存する。
    """
    try:
        # BytesIOオブジェクトを直接openpyxlに渡す
        wb = openpyxl.load_workbook(target_excel_bytes_io)
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"シート「{sheet_name}」が存在しません。")
        ws = wb[sheet_name]

        # ... (以降のハイライト処理ロジックは変更なし) ...
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, int) and cell.value in result_dict:
                    if "ルース" in result_dict[cell.value] and len(result_dict[cell.value]) == 1:
                        cell.font = Font(name="HGS明朝E", color="00FF00", size=14)
                    elif "ルース" in result_dict[cell.value] and len(result_dict[cell.value]) > 1:
                        cell.font = Font(name="HGS明朝E", color="FFFF00", size=14)
                    elif len(result_dict[cell.value]) >= 1:
                        cell.font = Font(name="HGS明朝E", color="FF0000", size=14)
        wb.save(output_path)
        print(f"処理完了：{output_path} にハイライト済みファイルを保存しました。")
    except Exception as e:
        print(f"Excelファイルのハイライト処理中にエラーが発生しました: {e}")
        raise

# --- run_brand_search_process 関数の修正 ---
async def run_brand_search_process(
    brand_list_excel_upload_file, # UploadFile オブジェクト
    pdf_upload_file,              # UploadFile オブジェクト
    target_excel_upload_file,     # UploadFile オブジェクト
    target_excel_filename: str,
    sheet_name: str
):
    PDF_TARGET_COLUMN_HEADER = "ブランド/モデル/型番/重量/状態等"
    PDF_TARGET_COLUMN_INDEX = 2
    BRAND_EXCEL_COLUMN_INDEX = 2
    BRAND_EXCEL_START_ROW = 2

    # UploadFile.file から内容を読み取り BytesIO に変換
    brand_list_excel_content = await brand_list_excel_upload_file.read()
    brand_list_excel_bytes_io = BytesIO(brand_list_excel_content)

    pdf_content = await pdf_upload_file.read()
    pdf_bytes_io = BytesIO(pdf_content)

    target_excel_content = await target_excel_upload_file.read()
    target_excel_bytes_io = BytesIO(target_excel_content)


    print("ブランドリストExcelの読み込みを開始...")
    brand_list_set = read_excel_to_set(
        brand_list_excel_bytes_io, # BytesIOオブジェクトを渡す
        column_index=BRAND_EXCEL_COLUMN_INDEX,
        start_row=BRAND_EXCEL_START_ROW
    )

    if not brand_list_set:
        print("Excelからブランドリストを読み込めませんでした。")
        raise ValueError("ブランドリストExcelからブランドが読み込めませんでした。ファイル内容を確認してください。")
    print(f"Excelから{len(brand_list_set)}件のブランド名を読み込み、正規化しました。")

    print("PDFからのブランド検索を開始...")
    result_dict = find_brands_in_pdf_plumber(
        pdf_bytes_io, # BytesIOオブジェクトを渡す
        brand_list_set,
        target_column_header=PDF_TARGET_COLUMN_HEADER,
        target_column_index=PDF_TARGET_COLUMN_INDEX
    )

    if result_dict:
        print("\n" + "="*15 + " 検索結果 " + "="*15)
        for page, brands in sorted(result_dict.items()):
            brands_str = ", ".join(sorted(list(brands)))
            print(f"  ページ {page}: {brands_str}")
        print(f"\n合計 {len(result_dict)} ページでブランドが見つかりました。")
        print("="*40)
    else:
        print("\n指定されたブランド名はPDF内のどのページにも見つかりませんでした。")

    temp_dir = tempfile.mkdtemp()
    base, ext = os.path.splitext(target_excel_filename)
    safe_filename = "".join(c if c.isalnum() or c in ('.', '_') else '_' for c in f"{base}_highlighted{ext}")
    output_excel_path = os.path.join(temp_dir, safe_filename)

    print(f"ハイライト処理を開始。出力先: {output_excel_path}")
    highlight_matching_cells(
        target_excel_bytes_io, # BytesIOオブジェクトを渡す
        result_dict,
        sheet_name,
        output_excel_path
    )
    
    return output_excel_path, result_dict, temp_dir